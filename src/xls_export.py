"""Excel export for new incoming requests (real and test)."""
from __future__ import annotations

import io
import os
import re
from datetime import datetime
from typing import Dict, List

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None
    Font = None
    PatternFill = None


def _is_department_assignment(employee_name: str) -> bool:
    """Check if employee_name is a department assignment (not a person).
    
    Returns True for departments, teams (προϊστάμενοι), etc.
    """
    if not employee_name:
        return False
    
    dept_keywords = ['ΤΜΗΜΑ', 'ΔΙΕΥΘΥΝΣΗ', 'Προϊστάμενοι']
    name_upper = str(employee_name).upper()
    return any(keyword in name_upper for keyword in dept_keywords)


def _format_protocol(rec: Dict) -> str:
    """Compose protocol string like "<case_id>(<protocol_number>)/DD-MM-YYYY"."""
    # case_id is outside, protocol_number is inside parentheses (same as CLI format)
    case_id = (rec.get("case_id") or "").strip()
    protocol = (rec.get("protocol_number") or "").strip()
    submitted_at = (rec.get("submitted_at") or "").strip()
    try:
        # Accept ISO or date-like strings
        dt = datetime.fromisoformat(submitted_at.replace(" ", "T")) if "-" in submitted_at else None
    except Exception:
        dt = None
    date_str = dt.strftime("%d-%m-%Y") if dt else ""
    if case_id and protocol and date_str:
        return f"{case_id}({protocol})/{date_str}"
    if case_id and date_str:
        return f"{case_id}/{date_str}"
    if protocol and date_str:
        return f"{protocol}/{date_str}"
    return case_id or protocol or date_str or ""


def _extract_related_case_key(related_case: str) -> str:
    """Extract 'YYYY/CASE_ID' from a related_case string.

    Example: 'Αίτημα 2025/717316 ΜΟΥΡΑΤΙΔΟΥ-128272645' → '2025/717316'
    """
    if not related_case:
        return ""
    match = re.search(r'(\d{4}/\d+)', related_case)
    return match.group(1) if match else ""


def _resolve_settled_info(rec: Dict, settled_by_case_id: Dict) -> Dict:
    """Resolve settled info for a record with 3-step priority:

    1. submission_year/case_id  (direct match)
    2. protocol_number          (fallback)
    3. supplement parent        (if document_category contains Συμπληρωματικά)

    Returns dict with 'settled_date' and 'assigned_employee' keys
    (empty strings if not settled).
    """
    result = {"settled_date": "", "assigned_employee": ""}

    case_id = str(rec.get("case_id", "")).strip()
    submission_year = str(rec.get("submission_year", "")).strip()

    # Step 1: submission_year/case_id
    if submission_year and case_id:
        lookup_key = f"{submission_year}/{case_id}"
        if lookup_key in settled_by_case_id:
            info = settled_by_case_id[lookup_key]
            result["settled_date"] = info.get("settled_date", "")
            result["assigned_employee"] = info.get("assigned_employee", "")
            return result

    # Step 2: protocol_number
    protocol_num = str(rec.get("protocol_number", "")).strip()
    if protocol_num and protocol_num in settled_by_case_id:
        info = settled_by_case_id[protocol_num]
        result["settled_date"] = info.get("settled_date", "")
        result["assigned_employee"] = info.get("assigned_employee", "")
        return result

    # Step 3: supplement parent
    doc_category = str(rec.get("document_category", ""))
    if "Συμπληρωματι" in doc_category:
        related_case = (rec.get("related_case") or "").strip()
        parent_key = _extract_related_case_key(related_case)
        if parent_key and parent_key in settled_by_case_id:
            info = settled_by_case_id[parent_key]
            result["settled_date"] = info.get("settled_date", "")
            result["assigned_employee"] = info.get("assigned_employee", "")
            return result

    return result


def _get_request_origin_label(rec: Dict) -> str:
    """Return request origin label for open-apps lists.

    - "Αίτηση" for normal requests
    - "Συμπληρωματικά στην YYYY/NNNNN" for supplement requests with parent key
    - "Συμπληρωματικά" for supplement requests without parsable parent key
    """
    doc_category = str(rec.get("document_category", ""))
    if "Συμπληρωματι" in doc_category:
        related_case = (rec.get("related_case") or "").strip()
        parent_key = _extract_related_case_key(related_case)
        if parent_key:
            return f"Συμπληρωματικά στην {parent_key}"
        return "Συμπληρωματικά"
    return "Αίτηση"


def _write_sheet(ws, rows: List[Dict], title: str, settled_by_case_id: Dict | None = None):
    """Write sheet with incoming requests.
    
    Columns: Α/Α, Δ/νση, Αρ. Πρωτοκόλλου, ΤΥΠΟΣ, Διαδικασία, Συναλλασσόμενος, Ανάθεση σε, Διεκπεραιωμένη
    """
    if settled_by_case_id is None:
        settled_by_case_id = {}
    
    headers = ["Α/Α", "Δ/νση", "Αρ. Πρωτοκόλλου", "ΤΥΠΟΣ", "Διαδικασία", "Συναλλασσόμενος", "Ανάθεση σε", "Διεκπεραιωμένη"]

    header_font = Font(bold=True) if Font else None
    header_fill = PatternFill("solid", fgColor="DAE8FC") if PatternFill else None

    def _parse_date(s: str | None):
        if not s:
            return None
        s = str(s).strip()
        try:
            return datetime.fromisoformat(s.replace(" ", "T"))
        except Exception:
            pass
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s[:10], fmt)
            except Exception:
                continue
        return None

    rows_sorted = sorted(
        rows,
        key=lambda rec: (
            _parse_date(rec.get("submitted_at")) or datetime.min,
            str(rec.get("protocol_number") or ""),
        ),
    )

    # Title row (row 1)
    title_cell = ws.cell(row=1, column=1, value=title)
    if header_font:
        title_cell.font = header_font

    # Header row (row 2)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        if header_font:
            cell.font = header_font
        if header_fill:
            cell.fill = header_fill

    # Data: 8 columns
    col_vals = [[], [], [], [], [], [], [], []]
    
    for idx, rec in enumerate(rows_sorted, start=1):
        # Ανάθεση σε (charge) - Priority: settled case charge > current incoming charge
        charge_info = rec.get("_charge", {})
        employee = charge_info.get("employee", "") if charge_info.get("charged") else ""
        
        # Resolve settled info (direct match, protocol fallback, supplement parent)
        settled_info = _resolve_settled_info(rec, settled_by_case_id)
        settled_date = settled_info["settled_date"]
        assigned_employee = settled_info["assigned_employee"]
        
        # Priority: Use assignment from settled cases (W001_P_FLD10), else use incoming charge
        if assigned_employee:
            employee = assigned_employee
        elif employee and _is_department_assignment(employee):
            # Filter out department assignments - only show personal assignments
            employee = ""  # Clear department assignment - only show personal ones
        
        col_vals[0].append(idx)  # Α/Α
        col_vals[1].append(rec.get("directory", ""))
        col_vals[2].append(_format_protocol(rec))
        
        # ΤΥΠΟΣ (column 4) - show related case if this is a supplement
        doc_type = rec.get("document_category", "")
        related_case = (rec.get("related_case") or "").strip()
        if "Συμπληρωματι" in str(doc_type) and related_case:
            case_number = _extract_related_case_key(related_case)
            if case_number:
                doc_type = f"{doc_type}: {case_number}"
        
        col_vals[3].append(doc_type)
        col_vals[4].append(rec.get("procedure", ""))
        col_vals[5].append(rec.get("party", ""))
        col_vals[6].append(employee)  # Assignment from settled cases or current charge
        col_vals[7].append(settled_date)

    # Column widths
    col_widths = [6, 100, 40, 42, 120, 60, 50, 25]  # Α/Α, Δ/νση, Protocol, ΤΥΠΟΣ, Διαδικασία, Party, Employee, Settled
    for idx in range(8):
        col_letter = chr(ord("A") + idx)
        ws.column_dimensions[col_letter].width = col_widths[idx]

    # Write data rows starting from row 3
    r = 3
    for i in range(len(rows_sorted)):
        for col_idx in range(8):
            ws.cell(row=r, column=col_idx + 1, value=col_vals[col_idx][i])
        r += 1
    
    # Freeze panes: freeze first 2 rows (title and header)
    ws.freeze_panes = "A3"


def print_open_apps_terminal(digest: Dict, monitor_instance=None) -> None:
    """Εκτυπώνει στο terminal αναφορά με όλες τις ΑΝΟΙΚΤΕΣ δοκιμαστικές αιτήσεις.
    
    Ανοικτή = δοκιμαστική (per classify_records) ΚΑΙ ΔΕΝ βρίσκεται στις διεκπεραιωμένες.
    Διαχωρισμός σε δύο ουρές:
      - AUTO-CLOSE: έχει ανατεθεί σε υπάλληλο (charged=True)
      - MANUAL:     δεν έχει ανάθεση (charged=False)
    """
    # Φόρτωση δεδομένων
    settled_by_case_id = _load_settled_cases(monitor_instance=monitor_instance)

    incoming = digest.get("incoming", {})
    try:
        from test_users import classify_records
        _real, test_rows = classify_records(incoming.get("records", []) or [])
    except Exception as exc:
        print(f"❌ Αποτυχία classify_records: {exc}")
        return

    # Κράτα μόνο ανοικτές (χωρίς ημερομηνία διεκπεραίωσης)
    open_tests = []
    for rec in test_rows:
        settled_info = _resolve_settled_info(rec, settled_by_case_id)
        if not settled_info["settled_date"]:
            open_tests.append(rec)

    # Διαχωρισμός σε auto-close / manual
    auto_close = [r for r in open_tests if r.get("_charge", {}).get("charged")]
    manual     = [r for r in open_tests if not r.get("_charge", {}).get("charged")]

    reason_labels = {
        'internal_user': 'Εσωτ. χρήστης',
        'test_user':     'Δοκ. χρήστης',
        'test_company':  'Δοκ. εταιρεία',
    }

    def _fmt_date(s):
        if not s:
            return ''
        try:
            dt = datetime.fromisoformat(str(s).replace(" ", "T"))
            return dt.strftime("%d-%m-%Y")
        except Exception:
            return str(s)[:10]

    def _print_group(rows, label, icon):
        print(f"\n{icon} {label} ({len(rows)} αιτήσεις)")
        if not rows:
            print("   (καμία)")
            return
        print(f"  {'Α/Α':<4} {'Case ID':<12} {'Ημ/νία Υποβ.':<14} {'Λόγος':<18} {'Είδος':<32} {'Διεύθυνση':<42} {'Αιτών / Party':<42} {'Ημ/νία Κλεισίματος'}")
        print("  " + "-" * 190)
        for idx, r in enumerate(rows, start=1):
            cid   = str(r.get("case_id", ""))
            date_ = _fmt_date(r.get("submitted_at", ""))
            reason = reason_labels.get(r.get("test_reason", ""), r.get("test_reason", ""))
            origin = _get_request_origin_label(r)[:31]
            direc  = (r.get("directory") or "")[:41]
            party  = (r.get("party") or "")[:42]
            close_date = ""  # Κενό προς το παρόν - θα συμπληρωθεί όταν κλείσει η αίτηση
            print(f"  {idx:<4} {cid:<12} {date_:<14} {reason:<18} {origin:<32} {direc:<42} {party:<42} {close_date}")

    total = len(open_tests)
    total_test = len(test_rows)
    total_settled_test = total_test - total

    print("\n" + "="*80)
    print(" 🧪 ΑΝΟΙΚΤΕΣ ΔΟΚΙΜΑΣΤΙΚΕΣ ΑΙΤΗΣΕΙΣ".center(80))
    print("="*80)
    print(f"  Σύνολο δοκιμαστικών (snapshot):   {total_test}")
    print(f"  Ήδη διεκπεραιωμένες (κλειστές):   {total_settled_test}")
    print(f"  Ανοικτές (χρειάζονται ενέργεια):  {total}")
    print(f"    ├─ ✅ Ανατεθειμένες (charged=True):  {len(auto_close)}")
    print(f"    └─ ⚠️  Χωρίς ανάθεση (charged=False): {len(manual)}")
    print("="*80)

    _print_group(auto_close, "Ανατεθειμένες δοκιμαστικές (charged=True — υποψήφιες για μελλοντικό αυτόματο κλείσιμο)", "✅")
    _print_group(manual,     "Χωρίς ανάθεση (charged=False — απαιτείται χειροκίνητη διαχείριση)", "⚠️")

    print("\n" + "="*80 + "\n")


def build_open_apps_xls(digest: Dict, monitor_instance=None, file_path: str | None = None) -> bytes | str:
    """Build an XLSX with two sheets for open test requests:
    - Sheet 1: Ανατεθειμένες (charged=True)
    - Sheet 2: Χωρίς Ανάθεση (charged=False)
    
    Columns: Α/Α, Case ID, Ημ/νία Υποβ., Λόγος, Διεύθυνση, Αιτών/Party, Ανάθεση σε, Ημ/νία Κλεισίματος
    """
    if not Workbook:
        raise RuntimeError("openpyxl not installed")

    # Reuse detection logic from print_open_tests_terminal()
    settled_by_case_id = _load_settled_cases(monitor_instance=monitor_instance)
    
    incoming = digest.get("incoming", {})
    try:
        from test_users import classify_records
        _real, test_rows = classify_records(incoming.get("records", []) or [])
    except Exception as exc:
        raise RuntimeError(f"Αποτυχία classify_records: {exc}")

    # Κράτα μόνο ανοικτές (χωρίς ημερομηνία διεκπεραίωσης)
    open_tests = []
    for rec in test_rows:
        settled_info = _resolve_settled_info(rec, settled_by_case_id)
        if not settled_info["settled_date"]:
            open_tests.append(rec)

    # Διαχωρισμός σε charged / uncharged
    charged = [r for r in open_tests if r.get("_charge", {}).get("charged")]
    uncharged = [r for r in open_tests if not r.get("_charge", {}).get("charged")]

    reason_labels = {
        'internal_user': 'Εσωτ. χρήστης',
        'test_user':     'Δοκ. χρήστης',
        'test_company':  'Δοκ. εταιρεία',
    }

    def _fmt_date(s):
        if not s:
            return ''
        try:
            dt = datetime.fromisoformat(str(s).replace(" ", "T"))
            return dt.strftime("%d-%m-%Y")
        except Exception:
            return str(s)[:10]

    def _parse_ddmmyyyy(s):
        if not s:
            return None
        try:
            return datetime.strptime(str(s).strip()[:10], "%d-%m-%Y")
        except Exception:
            return None

    def _current_rows_by_key(rows: List[Dict], group: str) -> Dict[str, Dict]:
        out = {}
        for rec in rows:
            key = str(rec.get("case_id", "")).strip()
            if not key:
                continue
            charge_info = rec.get("_charge", {})
            employee = charge_info.get("employee", "") if charge_info.get("charged") else ""
            out[key] = {
                "key": key,
                "submitted_at": _fmt_date(rec.get("submitted_at", "")),
                "reason": reason_labels.get(rec.get("test_reason", ""), rec.get("test_reason", "")),
                "origin": _get_request_origin_label(rec),
                "directory": rec.get("directory", ""),
                "party": rec.get("party", ""),
                "employee": employee,
                "close_date": "",
                "group": group,
            }
        return out

    def _load_existing_rows(path: str) -> Dict[str, Dict]:
        if load_workbook is None or not os.path.exists(path):
            return {}

        wb_existing = load_workbook(path)
        group_by_sheet = {
            "Ανατεθειμένες": "charged",
            "Χωρίς Ανάθεση": "uncharged",
        }
        loaded = {}

        for sheet_name, group in group_by_sheet.items():
            if sheet_name not in wb_existing.sheetnames:
                continue
            ws = wb_existing[sheet_name]
            for row in ws.iter_rows(min_row=3, values_only=True):
                key = str(row[1] or "").strip() if len(row) > 1 else ""
                if not key:
                    continue
                loaded[key] = {
                    "key": key,
                    "submitted_at": str(row[2] or "") if len(row) > 2 else "",
                    "reason": str(row[3] or "") if len(row) > 3 else "",
                    "origin": str(row[4] or "") if len(row) > 4 else "",
                    "directory": str(row[5] or "") if len(row) > 5 else "",
                    "party": str(row[6] or "") if len(row) > 6 else "",
                    "employee": str(row[7] or "") if len(row) > 7 else "",
                    "close_date": str(row[8] or "") if len(row) > 8 else "",
                    "group": group,
                }

        return loaded

    def _resolve_close_date_for_existing(row: Dict, default_close_date: str) -> str:
        existing_close = (row.get("close_date") or "").strip()
        if existing_close:
            return existing_close

        submitted_dt = _parse_ddmmyyyy(row.get("submitted_at", ""))
        key = (row.get("key") or "").strip()
        if submitted_dt and key:
            settled_key = f"{submitted_dt.year}/{key}"
            settled_info = settled_by_case_id.get(settled_key)
            if settled_info and settled_info.get("settled_date"):
                return settled_info.get("settled_date")

        return default_close_date

    def _write_open_tests_sheet(ws, rows: List[Dict], title: str):
        """Write sheet for open test requests."""
        headers = ["Α/Α", "Case ID", "Ημ/νία Υποβ.", "Λόγος", "Είδος", "Διεύθυνση", "Αιτών / Party", "Ανάθεση σε", "Ημ/νία Κλεισίματος"]
        
        header_font = Font(bold=True) if Font else None
        header_fill = PatternFill("solid", fgColor="FFF2CC") if PatternFill else None
        
        # Title row (row 1)
        title_cell = ws.cell(row=1, column=1, value=title)
        if header_font:
            title_cell.font = header_font
        
        # Header row (row 2)
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=c, value=h)
            if header_font:
                cell.font = header_font
            if header_fill:
                cell.fill = header_fill
        
        # Data: 9 columns
        col_vals = [[], [], [], [], [], [], [], [], []]

        rows_sorted = sorted(
            rows,
            key=lambda r: (
                _parse_ddmmyyyy(r.get("submitted_at", "")) or datetime.min,
                str(r.get("key") or ""),
            ),
        )

        for idx, rec in enumerate(rows_sorted, start=1):
            col_vals[0].append(idx)  # Α/Α
            col_vals[1].append(str(rec.get("key", "")))
            col_vals[2].append(rec.get("submitted_at", ""))
            col_vals[3].append(rec.get("reason", ""))
            col_vals[4].append(rec.get("origin", ""))
            col_vals[5].append(rec.get("directory", ""))
            col_vals[6].append(rec.get("party", ""))
            col_vals[7].append(rec.get("employee", ""))
            col_vals[8].append(rec.get("close_date", ""))
        
        # Write data starting from row 3
        for row_idx, *_ in enumerate(col_vals[0], start=3):
            for col_idx in range(9):
                ws.cell(row=row_idx, column=col_idx + 1, value=col_vals[col_idx][row_idx - 3])
        
        # Column widths: Α/Α(6), CaseID(12), Date(14), Reason(18), Type(34), Directory(50), Party(60), Assignment(40), CloseDate(18)
        col_widths = [6, 12, 14, 18, 34, 50, 60, 40, 18]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width
    
    current_charged = _current_rows_by_key(charged, group="charged")
    current_uncharged = _current_rows_by_key(uncharged, group="uncharged")
    current_all = {**current_charged, **current_uncharged}

    existing_all = _load_existing_rows(file_path) if file_path else {}
    close_date_today = datetime.now().strftime("%d-%m-%Y")

    merged_all = {}
    for key in set(existing_all.keys()) | set(current_all.keys()):
        if key in current_all:
            merged_all[key] = current_all[key]
            merged_all[key]["close_date"] = ""
        else:
            prev_row = existing_all[key]
            prev_row["close_date"] = _resolve_close_date_for_existing(prev_row, close_date_today)
            merged_all[key] = prev_row

    rows_charged = [r for r in merged_all.values() if r.get("group") == "charged"]
    rows_uncharged = [r for r in merged_all.values() if r.get("group") == "uncharged"]

    # Build workbook
    wb = Workbook()
    
    # Sheet 1: Ανατεθειμένες
    ws1 = wb.active
    ws1.title = "Ανατεθειμένες"
    _write_open_tests_sheet(ws1, rows_charged, "✅ Ανατεθειμένες δοκιμαστικές (charged=True)")
    
    # Sheet 2: Χωρίς Ανάθεση
    ws2 = wb.create_sheet("Χωρίς Ανάθεση")
    _write_open_tests_sheet(ws2, rows_uncharged, "⚠️ Χωρίς ανάθεση (charged=False)")
    
    # Return bytes or save to file
    if file_path:
        wb.save(file_path)
        return file_path
    
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_requests_xls(digest: Dict, scope: str = "new", file_path: str | None = None, monitor_instance = None) -> bytes | str:
    """Build an XLSX with two sheets (test, real).

    scope:
      - "new": only today's new requests (incoming.real_new / incoming.test_new)
      - "all": all requests from the snapshot (classified via test_users)
    
    monitor_instance: optional pre-authenticated PKMMonitor for loading settled cases

    Returns bytes if file_path is None; else writes to file_path and returns the path.
    """
    if Workbook is None:
        raise ImportError("openpyxl is not installed. Please 'pip install openpyxl'.")

    # Load settled cases for cross-reference
    settled_by_case_id = _load_settled_cases(monitor_instance=monitor_instance)

    incoming = digest.get("incoming", {})
    if scope == "all":
        try:
            from test_users import classify_records
            real_rows, test_rows = classify_records(incoming.get("records", []) or [])
        except Exception:
            # Fallback: treat all as real
            real_rows = incoming.get("records", []) or []
            test_rows = []
        title_test = "Δοκιμαστικές Αιτήσεις (Όλες)"
        title_real = "Πραγματικές Αιτήσεις (Όλες)"
    else:
        real_rows = incoming.get("real_new", []) or []
        test_rows = incoming.get("test_new", []) or []
        title_test = "Νέες Δοκιμαστικές Αιτήσεις"
        title_real = "Νέες Πραγματικές Αιτήσεις"

    wb = Workbook()
    ws_test = wb.active
    ws_test.title = "Δοκιμαστικές"
    ws_real = wb.create_sheet("Πραγματικές")

    _write_sheet(ws_test, test_rows, title_test, settled_by_case_id)
    _write_sheet(ws_real, real_rows, title_real, settled_by_case_id)

    if file_path:
        wb.save(file_path)
        return file_path

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _load_settled_cases(monitor_instance = None) -> Dict:
    """Load settled cases from queryId=19 using existing authenticated session.
    
    Args:
        monitor_instance: pre-authenticated PKMMonitor instance. If None, returns empty dict.
    
    Returns dict: {W001_P_FLD2: {'settled_date': 'DD-MM-YYYY', ...}, ...}
    """
    try:
        if monitor_instance is None:
            print("[DEBUG] No monitor instance provided for settled cases")
            return {}
        
        monitor = monitor_instance
        
        print("[DEBUG] Loaded settled cases from queryId=19")
        
        # Fetch settled cases (queryId=19)
        settled_params = {
            'isPoll': 'false',
            'queryId': '19',
            'queryOwner': '2',
            'isCase': 'false',
            'stateId': 'welcomeGrid-45_dashboard0',
            'page': '1',
            'start': '0',
            'limit': '500'
        }
        
        data = monitor.fetch_data(settled_params)
        if not data or not data.get('success'):
            print("[DEBUG] Failed to fetch settled cases")
            return {}
        
        settled_by_case_id = {}
        for rec in data.get('data', []):
            # W001_P_FLD2 = protocol number (like "2026/124212")
            protocol_num = str(rec.get('W001_P_FLD2', '')).strip()
            settled_date = str(rec.get('W001_P_FLD3', '')).strip()
            assigned_employee = str(rec.get('W001_P_FLD10', '')).strip()  # Employee assigned (charged) to the case
            
            if protocol_num:
                # Map by the protocol number directly (e.g., "2026/124212")
                settled_by_case_id[protocol_num] = {
                    'settled_date': settled_date,
                    'assigned_employee': assigned_employee
                }
        
        print(f"[DEBUG] Loaded {len(settled_by_case_id)} settled cases")
        return settled_by_case_id
    except Exception as e:
        # Silently fail to avoid breaking the export if settled cases unavailable
        print(f"[DEBUG] Error loading settled cases: {e}")
        return {}
