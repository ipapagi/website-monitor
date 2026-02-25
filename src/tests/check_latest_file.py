#!/usr/bin/env python
import sys
sys.path.insert(0, './src')

from openpyxl import load_workbook
from pathlib import Path

# Find 2026-02-02 files
report_path = Path('./data/outputs/reports')
xlsx_files = [f for f in report_path.glob('2026-02-02*.xlsx')]

if not xlsx_files:
    print("No 2026-02-02 files found")
    sys.exit(1)

# Sort by size to get largest
xlsx_files.sort(key=lambda x: x.stat().st_size, reverse=True)
test_file = xlsx_files[0]

print(f"📄 Testing: {test_file.name}\n")
print("=" * 100)

wb = load_workbook(test_file)
ws = wb['ΝΕΑ']

print("📋 CHECKING COLUMNS WITH DIRECTORY (after fix):\n")

# Check first 5 rows
for row in range(3, min(8, ws.max_row + 1)):
    print(f"\n📍 Row {row}:")
    h_val = ws[f'H{row}'].value  # Δ/νση
    k_val = ws[f'K{row}'].value  # Αρ. Πρωτ.
    
    print(f"  H (Δ/νση Πρωτοκόλλου):")
    print(f"    {h_val}")
    print(f"  K (Αρ. Πρωτ.):")
    print(f"    {k_val}")
    
    # Check for parentheses
    if h_val and '(' in str(h_val):
        print(f"  ❌ Column H STILL has parentheses!")
    else:
        print(f"  ✅ Column H clean (no parentheses)")
        
    if k_val and '(' in str(k_val):
        print(f"  ❌ Column K STILL has parentheses!")
    else:
        print(f"  ✅ Column K clean (no parentheses)")

print("\n" + "=" * 100)
