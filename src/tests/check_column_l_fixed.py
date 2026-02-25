#!/usr/bin/env python
import sys
sys.path.insert(0, './src')

from openpyxl import load_workbook
from pathlib import Path

# Find the first XLSX file generated
report_path = Path('./data/outputs/reports')
xlsx_files = list(report_path.glob('*.xlsx'))

if not xlsx_files:
    print("❌ Δεν βρέθηκαν αρχεία XLSX")
    sys.exit(1)

test_file = sorted(xlsx_files)[-1]  # Most recent
print(f"📄 Ελέγχω αρχείο: {test_file.name}\n")

wb = load_workbook(test_file)
ws_new = wb['ΝΕΑ']

# Check column L (protocol_date)
filled = 0
empty = 0
samples = []

for row_idx in range(3, min(15, ws_new.max_row + 1)):  # Check first 10 data rows
    protocol_date = ws_new[f'L{row_idx}'].value
    if protocol_date:
        filled += 1
        if len(samples) < 3:
            samples.append(f"  Row {row_idx}: {protocol_date}")
    else:
        empty += 1

print(f"📊 Στήλη L (Ημ/νία Πρωτ):")
print(f"  ✅ Γεμάτες: {filled}/10")
print(f"  ❌ Κενές: {empty}/10")

if samples:
    print(f"\n📋 Δείγματα:")
    for sample in samples:
        print(sample)
else:
    print(f"\n❌ ΔΕΝ υπάρχουν δεδομένα στη στήλη L")

# Also check column K for comparison
print(f"\n📊 Στήλη K (Αρ. Πρωτ.) - έλεγχος για σύγκριση:")
k_filled = 0
k_samples = []
for row_idx in range(3, min(15, ws_new.max_row + 1)):
    ar_prot = ws_new[f'K{row_idx}'].value
    if ar_prot:
        k_filled += 1
        if len(k_samples) < 2:
            k_samples.append(f"  Row {row_idx}: {ar_prot}")

print(f"  ✅ Γεμάτες: {k_filled}/10")
if k_samples:
    print(f"\n  📋 Δείγματα:")
    for sample in k_samples:
        print(sample)

print("\n✅ Έλεγχος ολοκληρώθηκε")
