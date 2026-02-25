#!/usr/bin/env python3
"""Check what's in the Excel file for 793923"""

import sys
import os

# Setup path for imports
from src_setup import *

from openpyxl import load_workbook
import json

# Load the Excel file
excel_file = 'Διαδικασίες - εισερχόμενες αιτήσεις.xlsx'

if not os.path.exists(excel_file):
    print(f"❌ File not found: {excel_file}")
    sys.exit(1)

print(f"\n📊 Reading: {excel_file}")
wb = load_workbook(excel_file)
ws = wb.active

print(f"\n🔍 Searching for: 793923")
print("="*80)

found_rows = []
for row_num, row in enumerate(ws.iter_rows(values_only=False), 1):
    if row_num == 1:  # Skip header
        continue
    
    # Check all cells in the row for the number
    row_data = []
    for cell in row:
        row_data.append(cell.value)
        if cell.value and str(cell.value).find('793923') != -1:
            found_rows.append((row_num, row_data))
            break

if not found_rows:
    print(f"❌ Not found in Excel")
else:
    print(f"✅ Found {len(found_rows)} row(s)\n")
    
    for row_num, row_data in found_rows:
        print(f"📋 Row {row_num}:")
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Get actual row with headers
        for row_obj in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=False):
            for col_num, (header, cell) in enumerate(zip(headers, row_obj), 1):
                print(f"   Col {col_num} ({header}): {cell.value}")
        print()
