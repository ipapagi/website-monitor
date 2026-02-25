#!/usr/bin/env python
"""Ελέγχει αν οι αναφορές περιέχουν μόνο αιτήσεις της σωστής εβδομάδας"""
from openpyxl import load_workbook
from datetime import datetime, timedelta
import glob

def get_week_boundaries(date_str):
    """Παίρνει μια ημερομηνία και επιστρέφει την προηγούμενη εβδομάδα (Δευτέρα-Κυριακή)"""
    given_date = datetime.strptime(date_str, "%Y-%m-%d")
    
    # Βρες την Δευτέρα της παρούσας εβδομάδας
    days_since_monday = given_date.weekday()  # Monday=0, Sunday=6
    monday_of_this_week = given_date - timedelta(days=days_since_monday)
    
    # Η προηγούμενη εβδομάδα: Δευτέρα ως Κυριακή
    sunday_of_prev_week = monday_of_this_week - timedelta(days=1)
    monday_of_prev_week = sunday_of_prev_week - timedelta(days=6)
    
    return (
        monday_of_prev_week.strftime("%Y-%m-%d"),
        sunday_of_prev_week.strftime("%Y-%m-%d")
    )

def check_week(date_str, print_details=True):
    """Ελέγχει αναφορές για μια συγκεκριμένη ημερομηνία"""
    
    monday_str, sunday_str = get_week_boundaries(date_str)
    monday_date = datetime.strptime(monday_str, '%Y-%m-%d').date()
    sunday_date = datetime.strptime(sunday_str, '%Y-%m-%d').date()
    
    pattern = f'data\\outputs\\reports\\{date_str}_*.xlsx'
    files = glob.glob(pattern)
    
    if not files:
        return None
    
    results = {}
    
    for file_path in sorted(files):
        file_name = file_path.split('\\')[-1]
        dir_name = file_name.split('_', 2)[2].replace('_Αναφορά Εισερχομένων Αιτήσεων ΣΗΔΕ.xlsx', '')
        
        try:
            wb = load_workbook(file_path)
            
            new_count = 0
            old_count = 0
            
            if 'ΝΕΑ' in wb.sheetnames:
                ws = wb['ΝΕΑ']
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
                    if row[2]:
                        new_count += 1
            
            if 'ΠΑΛΙΑ' in wb.sheetnames:
                ws = wb['ΠΑΛΙΑ']
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
                    if row[2]:
                        old_count += 1
            
            results[dir_name] = {'new': new_count, 'old': old_count}
        
        except Exception as e:
            pass
    
    return {'monday': monday_date, 'sunday': sunday_date, 'data': results}

# Ελέγχει δύο εβδομάδες
print("Συγfxριση αναφορών δύο διαδοχικών εβδομάδων\n")
print("=" * 80)

week1 = check_week('2026-02-02')
week2 = check_week('2026-02-09')

if week1:
    print(f"\nEbdoada 1: {week1['monday']} έως {week1['sunday']}")
    print("\nGeniki Dieuthynsi | NEA | PALIA | SYNOLO")
    print("-" * 50)
    
    week1_totals = {}
    for dir_name, counts in sorted(week1['data'].items()):
        total = counts['new'] + counts['old']
        week1_totals[dir_name] = {'new': counts['new'], 'old': counts['old'], 'total': total}
        dir_short = dir_name[:30] if len(dir_name) <= 30 else dir_name[:27] + "..."
        print(f"{dir_short:30} | {counts['new']:3} | {counts['old']:5} | {total:5}")
    
    week1_total_new = sum(d['new'] for d in week1['data'].values())
    week1_total_old = sum(d['old'] for d in week1['data'].values())
    print("-" * 50)
    print(f"{'SYNOLO':30} | {week1_total_new:3} | {week1_total_old:5} | {week1_total_new+week1_total_old:5}")

if week2:
    print(f"\n\nEbdoada 2: {week2['monday']} έως {week2['sunday']}")
    print("\nGeniki Dieuthynsi | NEA | PALIA | SYNOLO")
    print("-" * 50)
    
    week2_totals = {}
    for dir_name, counts in sorted(week2['data'].items()):
        total = counts['new'] + counts['old']
        week2_totals[dir_name] = {'new': counts['new'], 'old': counts['old'], 'total': total}
        dir_short = dir_name[:30] if len(dir_name) <= 30 else dir_name[:27] + "..."
        print(f"{dir_short:30} | {counts['new']:3} | {counts['old']:5} | {total:5}")
    
    week2_total_new = sum(d['new'] for d in week2['data'].values())
    week2_total_old = sum(d['old'] for d in week2['data'].values())
    print("-" * 50)
    print(f"{'SYNOLO':30} | {week2_total_new:3} | {week2_total_old:5} | {week2_total_new+week2_total_old:5}")

# Sygkrish
if week1 and week2:
    print("\n\n" + "=" * 80)
    print("SYGKRISH: EBDOMADA 2 vs EBDOMADA 1")
    print("=" * 80)
    
    all_ok = True
    
    for dir_name in sorted(set(week1['data'].keys()) | set(week2['data'].keys())):
        w1_old = week1['data'].get(dir_name, {}).get('old', 0)
        w1_new = week1['data'].get(dir_name, {}).get('new', 0)
        w2_old = week2['data'].get(dir_name, {}).get('old', 0)
        w2_new = week2['data'].get(dir_name, {}).get('new', 0)
        
        # Perimenom: w2_old >= w1_old + w1_new (ta palia ths ebdoadas 2 na periexoun toylahiston ta palia + nea ths ebdoadas 1)
        expected_min = w1_old + w1_new
        
        status = "OK" if w2_old >= expected_min else "PROBLEMA"
        dir_short = dir_name[:25] if len(dir_name) <= 25 else dir_name[:22] + "..."
        print(f"{status:8} | {dir_short:25} | Week1: old={w1_old:3}+new={w1_new:3}={expected_min:3} <= Week2: old={w2_old:3}")
        
        if w2_old < expected_min:
            all_ok = False
    
    print("\n" + "=" * 80)
    if all_ok:
        print("OK! To PALIA phyllo ths ebdoadas 2 periexei toylahiston ta dedomena ths ebdoadas 1")
    else:
        print("PROBLEMA! Kaporea dedomena leipoun apo to PALIA phyllo ths ebdoadas 2")
