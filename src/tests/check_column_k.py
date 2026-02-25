#!/usr/bin/env python3
"""Έλεγχος της στήλης K στο generated XLSX"""
from openpyxl import load_workbook
import os

reports_dir = r"C:\Develop\Office\Python\check_politis\website-monitor\data\outputs\reports"

# Παίρνουμε το πρώτο XLSX αρχείο
xlsx_files = [f for f in os.listdir(reports_dir) if f.endswith('.xlsx')]

if not xlsx_files:
    print("❌ Δεν βρέθηκαν XLSX αρχεία")
    exit(1)

file_path = os.path.join(reports_dir, xlsx_files[0])
print(f"📄 Έλεγχος αρχείου: {xlsx_files[0]}\n")

wb = load_workbook(file_path)

# Έλεγχος στο ΝΕΑ sheet
if 'ΝΕΑ' in wb.sheetnames:
    ws = wb['ΝΕΑ']
    print("=" * 80)
    print("ΣΤΗΛΗ K (Αρ. Πρωτ.) - ΝΕΑ ΑΙΤΗΣΕΙΣ")
    print("=" * 80)
    count = 0
    filled = 0
    for row in range(3, min(13, ws.max_row + 1)):  # Πρώτες 10 εγγραφές
        case_id = ws[f'B{row}'].value
        ar_prot = ws[f'K{row}'].value
        
        count += 1
        if ar_prot:
            filled += 1
        
        status = "✅" if ar_prot else "❌"
        print(f"{status} Σειρά {row}: Case_ID={case_id}, Αρ.Πρωτ.={ar_prot}")
    
    print(f"\n📊 Συμπληρωμένη: {filled}/{count}")

# Έλεγχος στο ΠΑΛΙΑ sheet
if 'ΠΑΛΙΑ' in wb.sheetnames:
    ws = wb['ΠΑΛΙΑ']
    print("\n" + "=" * 80)
    print("ΣΤΗΛΗ K (Αρ. Πρωτ.) - ΠΑΛΙΕΣ ΑΙΤΗΣΕΙΣ")
    print("=" * 80)
    count = 0
    filled = 0
    for row in range(3, min(13, ws.max_row + 1)):  # Πρώτες 10 εγγραφές
        case_id = ws[f'B{row}'].value
        ar_prot = ws[f'K{row}'].value
        
        count += 1
        if ar_prot:
            filled += 1
        
        status = "✅" if ar_prot else "❌"
        print(f"{status} Σειρά {row}: Case_ID={case_id}, Αρ.Πρωτ.={ar_prot}")
    
    print(f"\n📊 Συμπληρωμένη: {filled}/{count}")

print("\n✨ Έλεγχος ολοκληρώθηκε!")
