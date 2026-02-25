#!/usr/bin/env python
import sys
sys.path.insert(0, './src')

from openpyxl import load_workbook
from pathlib import Path

# Find the largest XLSX file
report_path = Path('./data/outputs/reports')
xlsx_files = list(report_path.glob('*.xlsx'))
xlsx_files.sort(key=lambda x: x.stat().st_size, reverse=True)
test_file = xlsx_files[0]

print(f"📄 Testing: {test_file.name}\n")
print("=" * 100)

wb = load_workbook(test_file)
ws = wb['ΝΕΑ']

print("📋 CHECKING COLUMNS WITH DIRECTORY/DEPARTMENT:\n")

# Check first 5 rows
for row in range(3, min(8, ws.max_row + 1)):
    print(f"\n📍 Row {row}:")
    h_val = ws[f'H{row}'].value  # Δ/νση Πρωτοκόλλου
    i_val = ws[f'I{row}'].value  # Τμήμα
    j_val = ws[f'J{row}'].value  # Αρ. Πρωτ Υπηρ
    k_val = ws[f'K{row}'].value  # Αρ. Πρωτ.
    
    print(f"  H (Δ/νση): {h_val}")
    print(f"  I (Τμήμα): {i_val}")
    print(f"  K (Αρ. Πρωτ.): {k_val}")
    
    # Check for parentheses
    if h_val and '(' in str(h_val):
        print(f"  ⚠️  Column H has parentheses!")
    if i_val and '(' in str(i_val):
        print(f"  ⚠️  Column I has parentheses!")
    if k_val and '(' in str(k_val):
        print(f"  ⚠️  Column K has parentheses!")

print("\n" + "=" * 100)
