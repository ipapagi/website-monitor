#!/usr/bin/env python3
"""Export charges to Excel for quick inspection"""

import sys
import os
from datetime import datetime

# Setup path
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(script_dir))
src_dir = os.path.join(project_root, 'src')
if src_dir not in sys.path:
    sys.path.insert(0, src_dir)

from monitor import PKMMonitor
from charges import fetch_charges_combined
from utils import load_config
from config import get_project_root

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERR openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)


def is_department(charge_text):
    """Check if charge is a department (group) instead of individual employee"""
    if not charge_text:
        return False
    
    # Departments contain these keywords
    dept_keywords = [
        'ΤΜΗΜΑ',
        'ΔΙΕΥΘΥΝΣΗ',
        'ΓΡΑΦΕΙΟ',
        'ΥΠΗΡΕΣΙΑ',
        'ΚΕΝΤΡΟ',
        '(Προϊστάμενοι)',
        'ΠΕ ',  # Position code that indicates department
    ]
    
    text = str(charge_text).upper()
    return any(keyword in text for keyword in dept_keywords)


def export_charges():
    """Export personal charges to Excel (excluding department assignments)"""
    
    # Load config
    root = get_project_root()
    cfg_path = os.path.join(root, 'config', 'config.yaml')
    config = load_config(cfg_path)
    
    # Connect
    monitor = PKMMonitor(
        base_url=config.get('base_url', 'https://shde.pkm.gov.gr'),
        urls=config.get('urls', {}),
        api_params=config.get('api_params', {}),
        login_params=config.get('login_params', {}),
        check_interval=config.get('check_interval', 300),
        username=config.get('username'),
        password=config.get('password'),
        session_cookies=config.get('session_cookies')
    )
    
    if not monitor.logged_in and not monitor.login():
        print("[ERR] Login failed")
        return 1
    
    print("\n[*] Fetching charges...")
    result = fetch_charges_combined(monitor)
    
    # fetch_charges_combined returns tuple (records, charges_dict)
    if isinstance(result, tuple):
        charges_records, charges_by_pkm = result
    else:
        # In case it's already a dict
        charges_by_pkm = result
        charges_records = []
    
    # Filter: keep only personal employee charges (exclude departments)
    personal_charges = {}
    for pkm, charge_data in charges_by_pkm.items():
        if isinstance(charge_data, dict):
            charge_text = charge_data.get('USER_GROUP_ID_TO', '')
        else:
            charge_text = str(charge_data) if charge_data else ''
        
        # Accept only if NOT a department
        if charge_text and not is_department(charge_text):
            personal_charges[pkm] = charge_data
    
    print("[*] Total in system: %d PKMs" % len(charges_by_pkm))
    print("[OK] Personal assignments (non-departments): %d" % len(personal_charges))
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Personal Charges"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["PKM", "Employee", "From", "Source", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    row = 2
    for pkm in sorted(personal_charges.keys()):
        charge_data = personal_charges[pkm]
        
        # charge_data is dict from queryId=3 with field USER_GROUP_ID_TO
        if isinstance(charge_data, dict):
            charge = charge_data.get('USER_GROUP_ID_TO', '')
            from_user = charge_data.get('USER_ID_FROM', '')
            source = 'queryId=3'
        else:
            charge = str(charge_data) if charge_data else ''
            from_user = ''
            source = ''
        
        ws.cell(row=row, column=1).value = pkm
        ws.cell(row=row, column=2).value = charge
        ws.cell(row=row, column=3).value = from_user
        ws.cell(row=row, column=4).value = source
        ws.cell(row=row, column=5).value = "✓" if charge else ""
        
        # Alternate row colors
        if row % 2 == 0:
            fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill
        
        row += 1
    
    # Adjust columns
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    
    # Save to data/outputs
    from pathlib import Path
    output_dir = Path(root) / "data" / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / f"Charges_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(output_file))
    
    print("\n[OK] Exported to: %s" % output_file)
    print("[*] Summary:")
    print("    Total in system: %d PKMs" % len(charges_by_pkm))
    print("    Personal assignments: %d" % len(personal_charges))
    if len(personal_charges) > 0:
        print("    Coverage: %.1f%%" % (100*len(personal_charges)/len(charges_by_pkm)))
    
    return 0


if __name__ == '__main__':
    sys.exit(export_charges() or 0)
