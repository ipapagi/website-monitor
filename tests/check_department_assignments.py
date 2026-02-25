#!/usr/bin/env python3
"""Check for department assignments in the Excel export."""

import sys
sys.path.insert(0, 'src')

from openpyxl import load_workbook
from tests.analyze_department_source import is_department

xlsx_path = r'data/Διαδικασίες - εισερχόμενες αιτήσεις.xlsx'
wb = load_workbook(xlsx_path)

print("=" * 80)
print("CHECKING FOR DEPARTMENT ASSIGNMENTS IN EXCEL")
print("=" * 80)

for sheet_name in ['Πραγματικές', 'Δοκιμαστικές']:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    print(f"\n📄 Sheet: {sheet_name}\n")
    
    department_rows = []
    personal_rows = []
    
    # Check each data row (starting from row 3)
    for row in range(3, ws.max_row + 1):
        assignment = ws.cell(row, 7).value  # Column 7 = Ανάθεση σε
        
        if assignment:  # If there's an assignment
            if is_department(assignment):
                department_rows.append((row, assignment))
            else:
                personal_rows.append((row, assignment))
    
    total_assigned = len(department_rows) + len(personal_rows)
    unassigned = ws.max_row - 2 - total_assigned
    
    print(f"  Total rows: {ws.max_row - 2}")
    print(f"  Assigned to DEPARTMENT: {len(department_rows)}")
    print(f"  Assigned to PERSON: {len(personal_rows)}")
    print(f"  Unassigned: {unassigned}")
    
    if department_rows:
        print(f"\n  ❌ DEPARTMENT ASSIGNMENTS (first 5):")
        for row_num, assignment in department_rows[:5]:
            print(f"    Row {row_num}: {assignment[:60]}...")
    
    if personal_rows:
        print(f"\n  ✅ PERSONAL ASSIGNMENTS (first 3):")
        for row_num, assignment in personal_rows[:3]:
            print(f"    Row {row_num}: {assignment[:60]}...")

print("\n" + "=" * 80)
print("❌ PROBLEM: Department assignments should be filtered out!")
print("=" * 80)
