#!/usr/bin/env python3
"""Check the actual data in the Excel file and verify records."""

import sys
import os
from openpyxl import load_workbook

# Load the Excel file
xlsx_path = r"c:\Develop\Office\Python\check_politis\website-monitor\data\Διαδικασίες - εισερχόμενες αιτήσεις.xlsx"

wb = load_workbook(xlsx_path)
ws = wb['Πραγματικές']

print("=== Checking Πραγματικές Sheet ===\n")

# Check first 10 data rows
for row_num in range(3, 13):
    cell_aa = ws.cell(row_num, 1).value    # Α/Α
    cell_protocol = ws.cell(row_num, 3).value  # Αρ. Πρωτοκόλλου (Column 3)
    cell_tipo = ws.cell(row_num, 4).value      # ΤΥΠΟΣ
    cell_diasd = ws.cell(row_num, 7).value     # Ανάθεση σε
    cell_settled = ws.cell(row_num, 8).value   # Διεκπεραιωμένη
    
    print(f"Row {row_num}: Α/Α={cell_aa}, Protocol={cell_protocol}, Type={cell_tipo[:15] if cell_tipo else None}..., Settled={cell_settled}")

print("\n=== Checking Δοκιμαστικές Sheet ===\n")

ws = wb['Δοκιμαστικές']

# Check first 10 data rows
for row_num in range(3, 13):
    cell_aa = ws.cell(row_num, 1).value    # Α/Α
    cell_protocol = ws.cell(row_num, 3).value  # Αρ. Πρωτοκόλλου
    cell_tipo = ws.cell(row_num, 4).value      # ΤΥΠΟΣ
    cell_diasd = ws.cell(row_num, 7).value     # Ανάθεση σε
    cell_settled = ws.cell(row_num, 8).value   # Διεκπεραιωμένη
    
    print(f"Row {row_num}: Α/Α={cell_aa}, Protocol={cell_protocol}, Type={cell_tipo[:15] if cell_tipo else None}..., Settled={cell_settled}")
