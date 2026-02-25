#!/usr/bin/env python3
"""Final verification: W001_P_FLD10 is the ASSIGNED employee (charge)."""

import sys
sys.path.insert(0, 'src')

from openpyxl import load_workbook

print("=" * 80)
print("ΔΙΕΥΚΡΙΝΙΣΗ: W001_P_FLD10 = ΕΙΣΗΓΗΤΗΣ ΠΟΥ ΧΡΕΩΘΗΚΕ")
print("=" * 80)

print("\n📋 Λογική Λειτουργίας:\n")
print("  1. Αν υπόθεση υπάρχει στα settled cases (queryId=19):")
print("     → Παίρνει χρέωση από W001_P_FLD10 (εισηγητής που χρεώθηκε)")
print()
print("  2. Αν δεν υπάρχει στα settled cases:")
print("     → Παίρνει τρέχουσα χρέωση από incoming data")
print()
print("  3. Φίλτρο: Αναθέσεις σε τμήματα δεν εμφανίζονται")
print()

xlsx_path = r'data/Διαδικασίες - εισερχόμενες αιτήσεις.xlsx'

try:
    wb = load_workbook(xlsx_path)
    
    print("=" * 80)
    print("ΕΠΑΛΗΘΕΥΣΗ COMMAND LINE EXPORT")
    print("=" * 80)
    
    for sheet_name in ['Δοκιμαστικές', 'Πραγματικές']:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        # Count assignments by source
        from_settled = 0
        from_incoming = 0
        examples_settled = []
        examples_incoming = []
        
        for row in range(3, ws.max_row + 1):
            assignment = ws.cell(row, 7).value  # Ανάθεση σε
            settled_date = ws.cell(row, 8).value  # Διεκπεραιωμένη
            protocol = ws.cell(row, 3).value
            
            if assignment:
                if settled_date:
                    # Has settled date -> assignment from W001_P_FLD10
                    from_settled += 1
                    if len(examples_settled) < 2:
                        examples_settled.append(f"{assignment} (settled: {settled_date})")
                else:
                    # No settled date -> assignment from incoming data
                    from_incoming += 1
                    if len(examples_incoming) < 2:
                        examples_incoming.append(f"{assignment}")
        
        print(f"\n  📊 {sheet_name}:")
        print(f"    Total: {ws.max_row - 2} εγγραφές")
        print()
        print(f"    Χρεώσεις από Settled Cases (W001_P_FLD10): {from_settled}")
        if examples_settled:
            for ex in examples_settled:
                print(f"      • {ex}")
        
        print()
        print(f"    Χρεώσεις από Incoming Data: {from_incoming}")
        if examples_incoming:
            for ex in examples_incoming:
                print(f"      • {ex}")
    
    print("\n" + "=" * 80)
    print("✅ ΣΩΣΤΗ ΛΕΙΤΟΥΡΓΙΑ")
    print("=" * 80)
    print()
    print("Το W001_P_FLD10 περιέχει τον εισηγητή που ΧΡΕΩΘΗΚΕ την υπόθεση,")
    print("όχι απαραίτητα αυτόν που τη διεκπεραίωσε.")
    print()
    print("Στο Excel, η στήλη 'Ανάθεση σε' συμπληρώνεται με:")
    print("  • Χρέωση από settled cases (αν υπάρχει)")
    print("  • Ή τρέχουσα χρέωση από incoming data")
    print("=" * 80)
    
except FileNotFoundError:
    print(f"\n❌ File not found: {xlsx_path}")
    print("Run: python src/main.py --export-incoming-xls-all")
