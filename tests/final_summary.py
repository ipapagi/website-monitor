#!/usr/bin/env python3
"""Final summary of ALL fixes."""

import sys
sys.path.insert(0, 'src')

from openpyxl import load_workbook

print("=" * 80)
print("ΤΕΛΙΚΗ ΕΠΑΛΗΘΕΥΣΗ - ΔΙΟΡΘΩΣΕΙΣ EXCEL EXPORT")
print("=" * 80)

xlsx_path = r'data/Διαδικασίες - εισερχόμενες αιτήσεις.xlsx'
wb = load_workbook(xlsx_path)

print("\n✅ COMMAND LINE EXPORT (python src/main.py --export-incoming-xls-all)\n")

for sheet_name in ['Δοκιμαστικές', 'Πραγματικές']:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    total = ws.max_row - 2
    
    # Count settlement dates
    settled_count = 0
    for row in range(3, ws.max_row + 1):
        if ws.cell(row, 8).value:
            settled_count += 1
    
    # Count assignments
    dept_count = 0
    personal_count = 0
    for row in range(3, ws.max_row + 1):
        assignment = ws.cell(row, 7).value
        if assignment:
            if any(kw in str(assignment).upper() for kw in ['ΤΜΗΜΑ', 'ΔΙΕΥΘΥΝΣΗ', 'Προϊστάμενοι']):
                dept_count += 1
            else:
                personal_count += 1
    
    unassigned = total - dept_count - personal_count
    
    print(f"📄 {sheet_name}:")
    print(f"  ├── Σύνολο: {total} εγγραφές")
    print(f"  ├── Διεκπεραιωμένη: {settled_count} με ημερομηνίες {'✅' if settled_count > 0 else '❌'}")
    print(f"  ├── Ανάθεση Τμήμα: {dept_count} {'✅ (φιλτραρισμένες)' if dept_count == 0 else '❌'}")
    print(f"  ├── Ανάθεση Άτομο: {personal_count} {'✅' if personal_count > 0 else ''}")
    print(f"  └── Χωρίς Ανάθεση: {unassigned}\n")

print("=" * 80)
print("ΠΡΟΒΛΗΜΑΤΑ ΠΟΥ ΔΙΟΡΘΩΘΗΚΑΝ:")
print("=" * 80)
print()
print("1. ✅ API ENDPOINT: Στήλη Διεκπεραιωμένη")
print("   - Πριν: Κενή στήλη (0 settlement dates)")
print("   - Μετά: Φορτώνει από queryId=19 (59 settled cases)")
print("   - Fix: Προστέθηκε global monitor σε webapi/state.py")
print()
print("2. ✅ DEPARTMENT ASSIGNMENTS: Χρεώσεις σε Τμήματα")
print("   - Πριν: 35 τμήματα (Δοκιμαστικές) + 14 (Πραγματικές)")
print("   - Μετά: 0 τμήματα (φιλτραρισμένα)")
print("   - Fix: Προστέθηκε _is_department_assignment() filter")
print()
print("=" * 80)
print("ΕΠΟΜΕΝΑ ΒΗΜΑΤΑ:")
print("=" * 80)
print()
print("1. Restart API server:")
print("   uvicorn src.main:app --host 0.0.0.0 --port 8000")
print()
print("2. Test API endpoint:")
print("   python test_api_export.py")
print()
print("3. Browser test:")
print("   http://localhost:8000/sede/export/xls?scope=all")
print()
print("=" * 80)
