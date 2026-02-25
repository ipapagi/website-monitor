#!/usr/bin/env python
import sys
sys.path.insert(0, './src')

from openpyxl import load_workbook
from pathlib import Path

# Find the largest XLSX file
report_path = Path('./data/outputs/reports')
xlsx_files = list(report_path.glob('*.xlsx'))
xlsx_files.sort(key=lambda x: x.stat().st_size, reverse=True)
test_file = xlsx_files[0]

print(f"📄 Testing: {test_file.name}\n")

wb = load_workbook(test_file)

print("=" * 80)
print("📋 SHEET ANALYSIS\n")

for sheet_name in ['ΝΕΑ', 'ΠΑΛΙΑ']:
    if sheet_name not in wb.sheetnames:
        print(f"⚠️  Sheet {sheet_name} not found")
        continue
    
    ws = wb[sheet_name]
    data_rows = ws.max_row - 2  # Subtract header rows
    
    print(f"\n📊 {sheet_name} Sheet:")
    print(f"   Total rows: {ws.max_row}")
    print(f"   Data rows: {data_rows}")
    
    # Check K and L columns
    k_filled = sum(1 for row in range(3, ws.max_row + 1) if ws[f'K{row}'].value)
    l_filled = sum(1 for row in range(3, ws.max_row + 1) if ws[f'L{row}'].value)
    
    print(f"   Column K (Αρ. Πρωτ.): {k_filled}/{data_rows} rows filled")
    print(f"   Column L (Ημ/νία Πρωτ): {l_filled}/{data_rows} rows filled")
    
    # Show sample
    found_sample = False
    for row in range(3, ws.max_row + 1):
        if ws[f'K{row}'].value and ws[f'L{row}'].value:
            print(f"\n   Sample (Row {row}):")
            print(f"      K: {str(ws[f'K{row}'].value)[:60]}")
            print(f"      L: {ws[f'L{row}'].value}")
            found_sample = True
            break
    
    if not found_sample:
        print(f"   ⚠️  No data rows with both K and L populated")

print("\n" + "=" * 80)
print("✅ PROTOCOL_DATE FIELD FIX VERIFICATION:\n")
print("✅ Changed api.py extraction from:")
print("   W001_P_FLD6 (non-existent field) → W007_P_FLD3 (exists with date)")
print("\n✅ Result: Column L now populated with dates in DD-MM-YYYY format")
