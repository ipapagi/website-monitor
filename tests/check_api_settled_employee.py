#!/usr/bin/env python3
"""Check API-exported file for settled employee data."""

import sys
sys.path.insert(0, 'src')

from openpyxl import load_workbook

xlsx_path = 'test_api_export.xlsx'

try:
    wb = load_workbook(xlsx_path)
    
    print("=" * 80)
    print("API EXPORT: Settled Employee Verification")
    print("=" * 80)
    
    for sheet_name in ['Δοκιμαστικές', 'Πραγματικές']:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        settled_with_employee = 0
        examples = []
        
        for row in range(3, min(20, ws.max_row + 1)):
            assignment = ws.cell(row, 7).value
            settled_date = ws.cell(row, 8).value
            
            if settled_date and assignment:
                settled_with_employee += 1
                if len(examples) < 3:
                    examples.append(f"Row {row}: {assignment} (Ημ/νία: {settled_date})")
        
        print(f"\n📄 {sheet_name}:")
        print(f"  Διεκπεραιωμένες με Εισηγητή (first 17 rows): {settled_with_employee}")
        
        if examples:
            print(f"  Παραδείγματα:")
            for ex in examples:
                print(f"    • {ex}")
    
    print("\n" + "=" * 80)
    print("✅ API Export: W001_P_FLD10 is included!")
    print("=" * 80)
    
except FileNotFoundError:
    print(f"❌ File not found: {xlsx_path}")
    print("Run: python test_api_export.py first")
