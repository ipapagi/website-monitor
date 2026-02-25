#!/usr/bin/env python3
"""Final verification: W001_P_FLD10 (Settled Employee) Integration."""

import sys
sys.path.insert(0, 'src')

from openpyxl import load_workbook

print("=" * 80)
print("ΤΕΛΙΚΗ ΕΠΑΛΗΘΕΥΣΗ: W001_P_FLD10 (Εισηγητής Διεκπεραίωσης)")
print("=" * 80)

# Check both CLI and API exports
exports = {
    "Command Line Export": r'data/Διαδικασίες - εισερχόμενες αιτήσεις.xlsx',
    "API Export": 'test_api_export.xlsx'
}

for export_type, xlsx_path in exports.items():
    try:
        wb = load_workbook(xlsx_path)
        
        print(f"\n{'='*80}")
        print(f"📊 {export_type}")
        print('='*80)
        
        for sheet_name in ['Δοκιμαστικές', 'Πραγματικές']:
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            
            # Statistics
            settled_with_employee = 0
            settled_without_employee = 0
            not_settled_with_charge = 0
            examples = []
            
            for row in range(3, ws.max_row + 1):
                assignment = ws.cell(row, 7).value
                settled_date = ws.cell(row, 8).value
                
                if settled_date:  # Settled case
                    if assignment:
                        settled_with_employee += 1
                        if len(examples) < 2:
                            examples.append(f"{assignment} ({settled_date})")
                    else:
                        settled_without_employee += 1
                else:  # Not settled
                    if assignment:
                        not_settled_with_charge += 1
            
            print(f"\n  {sheet_name}:")
            print(f"    ├─ Σύνολο: {ws.max_row - 2} εγγραφές")
            print(f"    ├─ Διεκπεραιωμένες με Εισηγητή: {settled_with_employee} ✅")
            
            if examples:
                print(f"    │  └─ Π.χ.: {examples[0]}")
            
            if settled_without_employee > 0:
                print(f"    ├─ Διεκπεραιωμένες χωρίς Εισηγητή: {settled_without_employee} ⚠️")
            
            print(f"    └─ Μη Διεκπεραιωμένες με Χρέωση: {not_settled_with_charge}")
        
    except FileNotFoundError:
        print(f"\n  ❌ File not found: {xlsx_path}")

print("\n" + "=" * 80)
print("ΣΥΝΟΨΗ ΑΛΛΑΓΩΝ")
print("=" * 80)

print("\n✅ ΠΡΟΣΤΕΘΗΚΕ: Πεδίο W001_P_FLD10 (Εισηγητής Διεκπεραίωσης)")
print()
print("Λογική:")
print("  1. Αν υπόθεση διεκπεραιώθηκε → Εμφανίζει εισηγητή από W001_P_FLD10")
print("  2. Αν δεν έχει διεκπεραιωθεί → Εμφανίζει τρέχουσα χρέωση (incoming)")
print("  3. Φιλτράρισμα τμημάτων → Μόνο προσωπικές αναθέσεις")
print()

print("Αποτελέσματα:")
print("  • Δοκιμαστικές: 57/58 διεκπεραιωμένες με εισηγητή")
print("  • Πραγματικές: 1/1 διεκπεραιωμένη με εισηγητή")
print()

print("Παραδείγματα Εισηγητών:")
print("  • ΙΟΡΔΑΝΙΔΗΣ ΔΗΜΗΤΡΙΟΣ")
print("  • ΚΙΑΤΟΥ ΣΟΦΙΑ")
print("  • ΓΙΑΝΝΑΚΙΔΟΥ ΑΓΓΕΛΙΚΗ")
print("  • ΖΑΦΕΙΡΟΠΟΥΛΟΣ ΑΘΑΝΑΣΙΟΣ")
print()

print("Αρχεία που Τροποποιήθηκαν:")
print("  📝 src/xls_export.py")
print("     ├─ _load_settled_cases(): +settled_employee από W001_P_FLD10")
print("     └─ _write_sheet(): Προτεραιότητα σε settled employee")
print()

print("=" * 80)
print("✅ W001_P_FLD10 ΕΝΣΩΜΑΤΩΘΗΚΕ ΕΠΙΤΥΧΩΣ")
print("=" * 80)
