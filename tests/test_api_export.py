#!/usr/bin/env python3
"""Download Excel from API and check settlement dates."""

import requests
import sys
from openpyxl import load_workbook

try:
    print("Downloading Excel from API...")
    # Bypass proxy for localhost
    proxies = {
        'http': None,
        'https': None,
    }
    resp = requests.get('http://localhost:8000/sede/export/xls?scope=all', timeout=60, proxies=proxies)
    print(f'Status: {resp.status_code}')
    print(f'Content-Type: {resp.headers.get("content-type")}')
    print(f'Size: {len(resp.content)} bytes')
    
    if resp.status_code == 200:
        with open('test_api_export.xlsx', 'wb') as f:
            f.write(resp.content)
        print('✅ Saved to test_api_export.xlsx')
        
        # Check the file
        print("\n=== Checking API-exported file ===\n")
        wb = load_workbook('test_api_export.xlsx')
        
        for sheet_name in ['Δοκιμαστικές', 'Πραγματικές']:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                total = ws.max_row - 2
                settled_count = 0
                
                for row in range(3, min(20, ws.max_row + 1)):
                    if ws.cell(row, 8).value:
                        settled_count += 1
                
                print(f"{sheet_name}:")
                print(f"  Total rows: {total}")
                print(f"  Rows with settlement dates (first 17): {settled_count}")
                print()
    else:
        print(f'Error: {resp.text[:500]}')
        sys.exit(1)
except Exception as e:
    import traceback
    print(f'Error: {e}')
    traceback.print_exc()
    sys.exit(1)
