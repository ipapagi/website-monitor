#!/usr/bin/env python
import sys
sys.path.insert(0, './src')

from openpyxl import load_workbook
from pathlib import Path

# Find the largest XLSX file generated
report_path = Path('./data/outputs/reports')
xlsx_files = list(report_path.glob('*.xlsx'))

if not xlsx_files:
    print("❌ Δεν βρέθηκαν αρχεία XLSX")
    sys.exit(1)

# Sort by file size
xlsx_files.sort(key=lambda x: x.stat().st_size, reverse=True)
test_file = xlsx_files[0]
print(f"📄 Μεγαλύτερο αρχείο: {test_file.name}\n")

wb = load_workbook(test_file)
ws_new = wb['ΝΕΑ']

print(f"Max row: {ws_new.max_row}")
print("\n=== DATA INSPECTION ===\n")

for row in range(3, min(8, ws_new.max_row + 1)):
    k_val = ws_new[f'K{row}'].value
    l_val = ws_new[f'L{row}'].value
    print(f"Row {row}:")
    print(f"  K (Αρ. Πρωτ.): {k_val}")
    print(f"  L (Ημ/νία Πρωτ): {l_val}")
    print()

# Count non-empty cells
k_count = sum(1 for row in range(3, ws_new.max_row + 1) if ws_new[f'K{row}'].value)
l_count = sum(1 for row in range(3, ws_new.max_row + 1) if ws_new[f'L{row}'].value)

print(f"📊 Summary:")
print(f"  Column K filled: {k_count} rows")
print(f"  Column L filled: {l_count} rows")
