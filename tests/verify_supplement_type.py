#!/usr/bin/env python3
"""Verify supplement documents show related case in ΤΥΠΟΣ column"""

from openpyxl import load_workbook

xlsx_path = r'data/Διαδικασίες - εισερχόμενες αιτήσεις.xlsx'

try:
    wb = load_workbook(xlsx_path)
    
    print("=" * 90)
    print("ΕΠΑΛΗΘΕΥΣΗ: Συμπληρωματικά Αιτήματα - Στήλη ΤΥΠΟΣ")
    print("=" * 90)
    
    for sheet_name in ['Δοκιμαστικές', 'Πραγματικές']:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        print(f"\n📊 {sheet_name}:")
        print("-" * 90)
        print(f"{'Αρ.':<4} {'ΤΥΠΟΣ':<45} {'Αρ. Πρωτοκόλλου':<35}")
        print("-" * 90)
        
        for row in range(3, min(ws.max_row + 1, 30)):
            row_num = row - 2
            doc_type = ws.cell(row, 4).value  # ΤΥΠΟΣ (column 4)
            protocol = ws.cell(row, 3).value  # Αρ. Πρωτοκόλλου (column 3)
            
            if doc_type and ('Συμπληρωματι' in str(doc_type) or 'Αίτηση' in str(doc_type)):
                doc_type_str = str(doc_type)[:43]
                protocol_str = str(protocol)[:33] if protocol else ""
                print(f"{row_num:<4} {doc_type_str:<45} {protocol_str:<35}")
        
        print("-" * 90)
    
except FileNotFoundError:
    print(f"❌ File not found: {xlsx_path}")
