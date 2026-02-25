#!/usr/bin/env python3
"""Verify the structure of the exported Excel file"""

from openpyxl import load_workbook

file_path = "data/Διαδικασίες - εισερχόμενες αιτήσεις.xlsx"

wb = load_workbook(file_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*80}")
    
    # Title (row 1)
    print(f"\nTitle (Row 1): {ws['A1'].value}")
    
    # Headers (row 2)
    print(f"\nHeaders (Row 2):")
    headers = []
    for col in range(1, 9):
        cell_value = ws.cell(row=2, column=col).value
        headers.append(cell_value)
        print(f"  Col {col}: {cell_value}")
    
    # Freeze panes
    print(f"\nFreeze Panes: {ws.freeze_panes}")
    
    # Sample data rows
    print(f"\nSample Data Rows (first 3):")
    for row_num in range(3, min(6, ws.max_row + 1)):
        print(f"\n  Row {row_num}:")
        for col_num, header in enumerate(headers, 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            print(f"    {header:20} = {cell_value}")
    
    print(f"\nTotal rows: {ws.max_row}")
