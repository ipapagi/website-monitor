#!/usr/bin/env python3
"""Verify W001_P_FLD10 (settled employee) is in Excel."""

import sys
sys.path.insert(0, 'src')

from openpyxl import load_workbook

xlsx_path = r'data/Διαδικασίες - εισερχόμενες αιτήσεις.xlsx'
wb = load_workbook(xlsx_path)

print("=" * 80)
print("VERIFICATION: W001_P_FLD10 (Settled Employee) in Excel")
print("=" * 80)

for sheet_name in ['Δοκιμαστικές', 'Πραγματικές']:
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    
    print(f"\n📄 Sheet: {sheet_name}\n")
    
    settled_with_employee = []
    settled_without_employee = []
    not_settled_with_charge = []
    not_settled_without_charge = []
    
    for row in range(3, ws.max_row + 1):
        protocol = ws.cell(row, 3).value  # Αρ. Πρωτοκόλλου
        assignment = ws.cell(row, 7).value  # Ανάθεση σε
        settled_date = ws.cell(row, 8).value  # Διεκπεραιωμένη
        
        if settled_date:  # Case is settled
            if assignment:
                settled_with_employee.append({
                    'row': row,
                    'protocol': protocol,
                    'employee': assignment,
                    'date': settled_date
                })
            else:
                settled_without_employee.append({
                    'row': row,
                    'protocol': protocol,
                    'date': settled_date
                })
        else:  # Case not settled
            if assignment:
                not_settled_with_charge.append({
                    'row': row,
                    'protocol': protocol,
                    'employee': assignment
                })
            else:
                not_settled_without_charge.append({
                    'row': row,
                    'protocol': protocol
                })
    
    print(f"  Διεκπεραιωμένες με Εισηγητή: {len(settled_with_employee)}")
    if settled_with_employee:
        print(f"    └─ Παραδείγματα (πρώτες 3):")
        for item in settled_with_employee[:3]:
            print(f"       Row {item['row']}: {item['employee']} (Ημ/νία: {item['date']})")
    
    print(f"\n  Διεκπεραιωμένες χωρίς Εισηγητή: {len(settled_without_employee)}")
    if settled_without_employee:
        print(f"    └─ Παραδείγματα:")
        for item in settled_without_employee[:2]:
            print(f"       Row {item['row']}: {item['protocol']} (Ημ/νία: {item['date']})")
    
    print(f"\n  Μη Διεκπεραιωμένες με Χρέωση: {len(not_settled_with_charge)}")
    if not_settled_with_charge:
        print(f"    └─ Παραδείγματα (πρώτες 2):")
        for item in not_settled_with_charge[:2]:
            print(f"       Row {item['row']}: {item['employee']}")
    
    print(f"\n  Μη Διεκπεραιωμένες χωρίς Χρέωση: {len(not_settled_without_charge)}")

print("\n" + "=" * 80)
print("ΣΥΝΟΨΗ")
print("=" * 80)
print()
print("✅ Διεκπεραιωμένες υποθέσεις πρέπει να έχουν εισηγητή από W001_P_FLD10")
print("✅ Μη διεκπεραιωμένες χρησιμοποιούν την τρέχουσα χρέωση")
print()
