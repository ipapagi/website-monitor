#!/usr/bin/env python3
"""Verify related cases are shown correctly in Excel"""

import sys
sys.path.insert(0, 'src')

from openpyxl import load_workbook

xlsx_path = r'data/Διαδικασίες - εισερχόμενες αιτήσεις.xlsx'

try:
    wb = load_workbook(xlsx_path)
    
    print("=" * 80)
    print("ΕΠΑΛΗΘΕΥΣΗ: Συμπληρωματικά Αιτήματα")
    print("=" * 80)
    
    for sheet_name in ['Δοκιμαστικές', 'Πραγματικές']:
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        print(f"\n📊 {sheet_name}:")
        print("-" * 80)
        print(f"{'Τύπος':<20} {'Αρ. Πρωτοκόλλου':<25} {'Σημείωση':<35}")
        print("-" * 80)
        
        supplements = 0
        normal = 0
        
        for row in range(3, min(ws.max_row + 1, 50)):  # Check first 50 rows
            doc_type = ws.cell(row, 4).value  # ΤΥΠΟΣ (column 4)
            protocol = ws.cell(row, 3).value  # Αρ. Πρωτοκόλλου (column 3)
            
            if doc_type and protocol:
                if 'Συμπληρωματι' in str(doc_type) or 'Συμπληρωμ' in str(doc_type):
                    supplements += 1
                    # Check if protocol contains '/' (year/case format)
                    is_related = '/' in str(protocol)
                    note = "✅ Σύνδεση με αρχική" if is_related else "⚠️ Δεν βρέθηκε σύνδεση"
                    print(f"{str(doc_type):<20} {str(protocol):<25} {note:<35}")
                elif 'Αίτηση' in str(doc_type):
                    normal += 1
        
        print("-" * 80)
        print(f"Σύνολο: Κανονικές={normal}, Συμπληρωματικά={supplements}")
    
except FileNotFoundError:
    print(f"❌ File not found: {xlsx_path}")
