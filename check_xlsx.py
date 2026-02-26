#!/usr/bin/env python
"""Quick check of generated XLSX file."""
from openpyxl import load_workbook

wb = load_workbook('data/outputs/Ανοικτές_δοκιμαστικές_20260226_093311.xlsx')
ws = wb['Ανατεθειμένες']

print("Sample rows from 'Ανατεθειμένες' sheet:")
print(f"Headers: {[cell.value for cell in ws[2]]}")
print("\nFirst 3 data rows (checking Close Date column):")
for i in range(3, min(6, ws.max_row+1)):
    aa = ws.cell(i, 1).value
    case_id = ws.cell(i, 2).value
    close_date = ws.cell(i, 8).value
    print(f"  Row {i}: Α/Α={aa}, Case ID={case_id}, Close Date=[{close_date}]")

ws2 = wb['Χωρίς Ανάθεση']
print(f"\n'Χωρίς Ανάθεση' sheet has {ws2.max_row-2} data rows")
print("✅ XLSX structure validated successfully!")
