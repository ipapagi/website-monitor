#!/usr/bin/env python
import sys
sys.path.insert(0, './src')

from openpyxl import load_workbook
from pathlib import Path

# Find the largest XLSX file
report_path = Path('./data/outputs/reports')
xlsx_files = list(report_path.glob('*.xlsx'))
xlsx_files.sort(key=lambda x: x.stat().st_size, reverse=True)
test_file = xlsx_files[0]

print(f"📄 Testing: {test_file.name}\n")
print("=" * 80)

wb = load_workbook(test_file)
ws_new = wb['ΝΕΑ']

# Column headers
headers = {
    'A': 'Α/α',
    'B': 'Αρ. Υπόθεσης',
    'C': 'Ημ/νία Έξόδου',
    'D': 'Κατηγορία Εγγράφου',
    'E': 'Συναλλασσόμενος',
    'F': 'Θέμα',
    'G': 'Ανάθεση σε',
    'H': 'Δ/νση',
    'I': 'Τμήμα',
    'J': 'Αρ. Πρωτ Υπηρ',
    'K': 'Αρ. Πρωτ.',
    'L': 'Ημ/νία Πρωτ'
}

# Count filled columns for first 5 data rows
print("✅ COLUMN POPULATION CHECK\n")
for row in range(3, min(8, ws_new.max_row + 1)):
    print(f"📍 Row {row}:")
    filled_cols = []
    for col in 'ABCDEFGHIJKL':
        val = ws_new[f'{col}{row}'].value
        if val:
            filled_cols.append(col)
            # Show sample of first 50 chars
            val_str = str(val)[:50] if val else "None"
            print(f"    {col}: {val_str}")
        else:
            print(f"    {col}: [EMPTY]")
    print()

# Summary statistics
print("=" * 80)
print("📊 COLUMN FILL STATISTICS:\n")
col_stats = {}
for col in 'ABCDEFGHIJKL':
    filled = sum(1 for row in range(3, ws_new.max_row + 1) if ws_new[f'{col}{row}'].value)
    total = ws_new.max_row - 2
    col_stats[col] = (filled, total)
    status = "✅" if filled == total else "⚠️" if filled > 0 else "❌"
    print(f"{status} Column {col} ({headers[col]:20}): {filled:3}/{total} rows filled")

all_filled = all(filled == total for filled, total in col_stats.values())
print("\n" + "=" * 80)
if all_filled:
    print("✅ ALL COLUMNS FULLY POPULATED!")
else:
    print("⚠️  Some columns are not fully populated")
